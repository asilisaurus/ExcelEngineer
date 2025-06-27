# ВАЖНО: Перед этой строкой не должно быть никаких пробелов или отступов.
# Код должен начинаться с самого начала строки.

import pandas as pd
import numpy as np
import os
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Игнорируем предупреждения для чистого вывода
warnings.filterwarnings('ignore', category=UserWarning, module='xlsxwriter')
warnings.filterwarnings('ignore', category=FutureWarning)


def clean_views(value):
    """
    Преобразует значение просмотров в числовой формат.
    Возвращает np.nan для некорректных значений.
    """
    if pd.isna(value) or str(value).strip() in ["", "Нет данных"]:
        return np.nan
    try:
        cleaned = str(value).replace(" ", "").replace(",", "").replace("'", "")
        return float(cleaned) if cleaned.replace('.', '', 1).isdigit() else np.nan
    except (ValueError, TypeError):
        return np.nan

def convert_to_report(source_path, output_path):
    """
    Основная функция для создания отчета.
    Использует чистый openpyxl для создания и форматирования файла,
    чтобы избежать проблем с зависимостями и повреждением файла.
    """
    # 1. ЗАГРУЗКА И ОБРАБОТКА ДАННЫХ
    xl_file = pd.ExcelFile(source_path)
    months = ["Янв25", "Фев25", "Мар25", "Март25", "Апр25", "Май25", "Июн25", 
              "Июл25", "Авг25", "Сен25", "Окт25", "Ноя25", "Дек25"]
    sheet_name = next((m for m in months if m in xl_file.sheet_names), None)
    
    if not sheet_name:
        raise ValueError(f"Лист с данными месяца не найден. Доступные листы: {', '.join(xl_file.sheet_names)}")
    
    print(f"Найден лист с данными: {sheet_name}")
    
    src = pd.read_excel(source_path, sheet_name=sheet_name, header=None)
    
    # Извлекаем данные по фиксированным диапазонам согласно структуре файла
    reviews_otz = src.iloc[6:15, [1, 3, 4, 6, 7, 10, 16, 13]].copy()
    reviews_apt = src.iloc[15:28, [1, 3, 4, 6, 7, 10, 16, 13]].copy()
    top20 = src.iloc[31:51, [1, 3, 4, 6, 7, 10, 16, 13]].copy()
    
    # ИСПРАВЛЕНИЕ: Активные обсуждения НЕ находятся в строках 52+
    # Строки начиная с 52 содержат "Комментарии в обсуждениях" которые не должны быть в отчете
    # Создаем пустой DataFrame для активных обсуждений
    active = pd.DataFrame()
    print("Активные обсуждения исключены из отчета согласно требованиям")

    columns = ["Площадка", "Тема", "Текст сообщения", "Дата", "Ник", "Просмотры", "Вовлечение", "Тип поста"]
    
    # Process only non-empty DataFrames
    dataframes_to_process = []
    if not reviews_otz.empty:
        dataframes_to_process.append(('reviews_otz', reviews_otz))
    if not reviews_apt.empty:
        dataframes_to_process.append(('reviews_apt', reviews_apt))
    if not top20.empty:
        dataframes_to_process.append(('top20', top20))
    if not active.empty:
        dataframes_to_process.append(('active', active))
    
    # Apply columns and clean data for each non-empty DataFrame
    for name, df in dataframes_to_process:
        df.columns = columns
        df['Просмотры'] = df['Просмотры'].apply(clean_views)
        # Заполняем пустые значения только в текстовых столбцах, чтобы избежать ошибок
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].fillna('')

    reviews = pd.concat([reviews_otz, reviews_apt], ignore_index=True)

    # 2. СОЗДАНИЕ И ФОРМАТИРОВАНИЕ EXCEL ФАЙЛА
    wb = Workbook()
    ws = wb.active
    
    month_name_map = {"Янв": "Январь", "Фев": "Февраль", "Мар": "Март", "Март": "Март", "Апр": "Апрель", "Май": "Май", "Июн": "Июнь", "Июл": "Июль", "Авг": "Август", "Сен": "Сентябрь", "Окт": "Октябрь", "Ноя": "Ноябрь", "Дек": "Декабрь"}
    # Handle both "Мар25" and "Март25" formats
    sheet_prefix = sheet_name[:4] if sheet_name.startswith("Март") else sheet_name[:3]
    report_month_name = month_name_map.get(sheet_prefix, "Отчет")
    ws.title = f"{report_month_name} 20{sheet_name[-2:]}"

    # --- Стили ---
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="2D1341", end_color="2D1341", fill_type="solid")
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="9CC2E5", end_color="9CC2E5", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_align_top = Alignment(horizontal='center', vertical='top', wrap_text=True)
    formula_header_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    formula_font = Font(bold=True)
    
    # --- Шапка отчета ---
    # Create month mapping with correct indices
    month_indices = {"Янв25": 1, "Фев25": 2, "Мар25": 3, "Март25": 3, "Апр25": 4, "Май25": 5, "Июн25": 6, 
                     "Июл25": 7, "Авг25": 8, "Сен25": 9, "Окт25": 10, "Ноя25": 11, "Дек25": 12}
    month_num = month_indices.get(sheet_name, 1)
    # ИСПРАВЛЕНИЕ: Формат даты точно как в образце - только дата без времени
    date_str = f"01.{month_num:02d}.2025"

    # Create proper header structure with merged cells
    ws.merge_cells('A1:B1'); ws['A1'] = 'Продукт'
    ws.merge_cells('C1:G1'); ws['C1'] = 'Акрихин - Фортедетрим'
    
    ws.merge_cells('A2:B2'); ws['A2'] = 'Период'  
    ws.merge_cells('C2:G2'); ws['C2'] = date_str
    
    ws.merge_cells('A3:B3'); ws['A3'] = 'План'
    ws.merge_cells('C3:G3'); ws['C3'] = 'Отзывы - 22, Комментарии - 650'

    # Apply formatting to header area (A1:H3)
    for row in range(1, 4):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = center_align

    # --- Заголовки таблицы ---
    main_headers = ["Площадка", "Тема", "Текст сообщения", "Дата", "Ник", "Просмотры", "Вовлечение", "Тип поста"]
    for col, header in enumerate(main_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(name='Arial', size=9, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_align
        
    table_header_fill = header_fill  # Use same fill as main header

    # --- Function to write dataframe data ---
    def write_df(ws, df, start_row):
        # Handle empty DataFrames
        if df.empty:
            return start_row
            
        for r_idx, (_, row) in enumerate(df.iterrows(), start_row):
            ws_row = r_idx
            
            # Handle each column - only replace "0" with "Нет данных", keep empty cells empty
            area = row.get('Площадка', '') if pd.notna(row.get('Площадка')) else ''
            topic = row.get('Тема', '') if pd.notna(row.get('Тема')) else ''
            message = row.get('Текст сообщения', '') if pd.notna(row.get('Текст сообщения')) else ''
            date_val = row.get('Дата') if pd.notna(row.get('Дата')) else None
            nick = row.get('Ник', '') if pd.notna(row.get('Ник')) else ''
            views = clean_views(str(row.get('Просмотры', 0)))
            engagement = row.get('Вовлечение', '') if pd.notna(row.get('Вовлечение')) else ''
            post_type = row.get('Тип поста', '') if pd.notna(row.get('Тип поста')) else ''
            
            # Replace zero, empty, or invalid values in views with "Нет данных"
            if views is np.nan or views == 0 or str(views) in ['0', '', 'nan', 'None']:
                views = 'Нет данных'
            
            # Set cell values
            ws[f'A{ws_row}'] = area
            ws[f'B{ws_row}'] = topic
            ws[f'C{ws_row}'] = message
            ws[f'D{ws_row}'] = date_val
            ws[f'E{ws_row}'] = nick
            ws[f'F{ws_row}'] = views
            ws[f'G{ws_row}'] = engagement
            ws[f'H{ws_row}'] = post_type
            
            # Apply formatting to each cell (no borders) with compact height
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                cell = ws[f'{col}{ws_row}']
                cell.font = Font(name='Arial', size=9)
                if col == 'D':
                    cell.number_format = 'DD.MM.YYYY'
                if col == 'F':
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align_top
            
            # Set row height to be more compact (like in the reference)
            ws.row_dimensions[ws_row].height = 12
                     
        return start_row + len(df)

    # Set default column widths for better appearance
    ws.column_dimensions['A'].width = 25  # Площадка
    ws.column_dimensions['B'].width = 20  # Тема
    ws.column_dimensions['C'].width = 50  # Текст сообщения
    ws.column_dimensions['D'].width = 12  # Дата
    ws.column_dimensions['E'].width = 15  # Ник
    ws.column_dimensions['F'].width = 12  # Просмотры
    ws.column_dimensions['G'].width = 12  # Вовлечение
    ws.column_dimensions['H'].width = 12  # Тип поста

    # --- ИСПРАВЛЕНИЕ: Правильная запись данных с синими разделителями ПЕРЕД секциями ---
    row_num = 5
    
    # Add blue divider BEFORE reviews section
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws[f'A{row_num}'] = "Отзывы"
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.alignment = center_align
    ws.row_dimensions[row_num].height = 12
    
    # Write reviews data AFTER the header
    row_num += 1
    start_reviews = row_num
    row_num = write_df(ws, reviews, row_num)
    
    # Add blue divider BEFORE top20 section
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws[f'A{row_num}'] = "Комментарии Топ-20 выдачи"
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.alignment = center_align
    ws.row_dimensions[row_num].height = 12
    
    # Write top20 data AFTER the header
    row_num += 1
    start_top20 = row_num
    row_num = write_df(ws, top20, row_num)
    
    # Add blue divider BEFORE active discussions
    row_num += 1
    ws.merge_cells(f'A{row_num}:H{row_num}')
    ws[f'A{row_num}'] = "Активные обсуждения (мониторинг)"
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.alignment = center_align
    ws.row_dimensions[row_num].height = 12
    
    # Write active data AFTER the header
    row_num += 1
    start_active = row_num
    row_num = write_df(ws, active, row_num)
    
    last_data_row = row_num - 1

    # --- ИСПРАВЛЕНИЕ РАСЧЕТОВ: Итоговые формулы с точными значениями ---
    formula_start_row = last_data_row + 3
    
    # Summary calculations matching the expected format with CORRECTED calculations
    ws.merge_cells(f'A{formula_start_row}:E{formula_start_row}')
    ws[f'A{formula_start_row}'] = "Суммарное количество просмотров"
    
    # ИСПРАВЛЕНИЕ: Calculate the actual sum value correctly - only numeric values, excluding "Нет данных"
    total_views = 0
    # Combine all data sections that are not empty
    dfs_to_concat = []
    if not reviews_otz.empty:
        dfs_to_concat.append(reviews_otz)
    if not reviews_apt.empty:
        dfs_to_concat.append(reviews_apt)
    if not top20.empty:
        dfs_to_concat.append(top20)
    if not active.empty:
        dfs_to_concat.append(active)
    
    if dfs_to_concat:
        all_data = pd.concat(dfs_to_concat, ignore_index=True)
        for _, row in all_data.iterrows():
            views_value = clean_views(str(row.get('Просмотры', 0)))
            if views_value is not np.nan and isinstance(views_value, (int, float)) and views_value > 0:
                total_views += int(views_value)
    
    # Apply additional formatting to formula cells
    for col in range(1, 6):
        cell = ws.cell(row=formula_start_row, column=col)
        cell.fill = formula_header_fill
        cell.font = formula_font
        cell.alignment = left_align_top
    
    ws[f'F{formula_start_row}'] = total_views
    ws[f'F{formula_start_row}'].fill = formula_header_fill
    ws[f'F{formula_start_row}'].font = formula_font
    ws[f'F{formula_start_row}'].alignment = center_align
    print(f"Calculated total views: {total_views}")
    
    ws.merge_cells(f'A{formula_start_row+1}:E{formula_start_row+1}')
    ws[f'A{formula_start_row+1}'] = "Количество карточек товара (отзывы)"
    # Count total reviews from both OTZ and APT
    total_reviews_count = 0
    if not reviews_otz.empty:
        total_reviews_count += len(reviews_otz)
    if not reviews_apt.empty:
        total_reviews_count += len(reviews_apt)
    ws[f'F{formula_start_row+1}'] = total_reviews_count
    
    # Apply formatting to second formula row
    for col in range(1, 6):
        cell = ws.cell(row=formula_start_row+1, column=col)
        cell.fill = formula_header_fill
        cell.font = formula_font
        cell.alignment = left_align_top
    ws[f'F{formula_start_row+1}'].fill = formula_header_fill
    ws[f'F{formula_start_row+1}'].font = formula_font
    ws[f'F{formula_start_row+1}'].alignment = center_align
    
    ws.merge_cells(f'A{formula_start_row+2}:E{formula_start_row+2}')
    ws[f'A{formula_start_row+2}'] = "Количество обсуждений (форумы, сообщества, комментарии к статьям)"
    total_discussions = len(top20) + len(active)
    ws[f'F{formula_start_row+2}'] = total_discussions
    
    # Apply formatting to third formula row
    for col in range(1, 6):
        cell = ws.cell(row=formula_start_row+2, column=col)
        cell.fill = formula_header_fill
        cell.font = formula_font
        cell.alignment = left_align_top
    ws[f'F{formula_start_row+2}'].fill = formula_header_fill
    ws[f'F{formula_start_row+2}'].font = formula_font
    ws[f'F{formula_start_row+2}'].alignment = center_align
    
    ws.merge_cells(f'A{formula_start_row+3}:E{formula_start_row+3}')
    ws[f'A{formula_start_row+3}'] = "Доля обсуждений с вовлечением в диалог"
    
    # ИСПРАВЛЕНИЕ: Calculate engagement percentage correctly
    discussions_data = pd.concat([top20, active], ignore_index=True)
    engagement_count = sum(1 for _, row in discussions_data.iterrows() if row.get('Вовлечение') == 'есть')
    engagement_pct = engagement_count / total_discussions if total_discussions > 0 else 0
    ws[f'F{formula_start_row+3}'] = engagement_pct
    ws[f'F{formula_start_row+3}'].number_format = '0%'
    
    # Apply formatting to fourth formula row
    for col in range(1, 6):
        cell = ws.cell(row=formula_start_row+3, column=col)
        cell.fill = formula_header_fill
        cell.font = formula_font
        cell.alignment = left_align_top
    ws[f'F{formula_start_row+3}'].fill = formula_header_fill
    ws[f'F{formula_start_row+3}'].font = formula_font
    ws[f'F{formula_start_row+3}'].alignment = center_align
    
    # Set row heights for formula section to be compact
    for i in range(4):
        ws.row_dimensions[formula_start_row + i].height = 12
    
    # Add spacing and footnote
    footnote_row = formula_start_row + 6
    ws.merge_cells(f'A{footnote_row}:F{footnote_row}')
    ws[f'A{footnote_row}'] = "*Фин учета количества просмотров за прошедший период - суммирование 3х месяцев, спонтанных за публикацией"
    ws[f'A{footnote_row}'].font = Font(name='Arial', size=8, italic=True)
    ws[f'A{footnote_row}'].alignment = left_align_top
    ws.row_dimensions[footnote_row].height = 12

    # --- СОХРАНЕНИЕ ФАЙЛА ---
    try:
        wb.save(output_path)
        print(f"Отчет успешно создан: {output_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
        raise


if __name__ == "__main__":
    # Для тестирования
    import sys
    if len(sys.argv) >= 3:
        source_file = sys.argv[1]
        output_file = sys.argv[2]
        convert_to_report(source_file, output_file)
    else:
        print("Использование: python excel_processor.py <источник.xlsx> <результат.xlsx>")
